#-----------------------------------------------------------------------------------------------------------------------------
#Context: code to get stock screener based on the data from screener.in and predefined conditions
#Version: 2.0
#Date: October 13, 2021
#Input: screener file from screener.in(allstocklist.csv) 
#Output: priority categorization of stocks
#Author: pinakin.parkhe@gmail.com
#-----------------------------------------------------------------------------------------------------------------------------
from collections import defaultdict
from functools import reduce
from openpyxl import Workbook, load_workbook,styles
from openpyxl.styles import PatternFill
import time
import glob
import os
import json
import numpy as np
import pandas as pd
#-----------------------------------------------------------------------------------------------------------------------------
pd.set_option('display.max_colwidth', 500)
#-----------------------------------------------------------------------------------------------------------------------------
def getCondScreener(df_cond, condName, colIndex):
	'''
		Input:
			screener data, condition name, column index of condition
		Output:
			dataframe with conditions satsified, list of stocks where condition is not satsified,
			list of stocks where condition is satisfied
		Context:
			function for condition screening:
				1. MCAP > 500 
				2. PE < 9
				3. PB < 6
				4. ROE > 15
				5. ROCE > 15
				6. DY > 0.5
				7. CR > 2
				8. DE < 1
	'''
	df_cond1 = {}
	listColCOND1 = []
	listColCOND0 = []
	for k in range(df_cond.shape[0]):
		if condName == 'MCAP':
			df_cond1[df_cond.iat[k,0]] = bool(df_cond.iat[k,colIndex] > 500)
		if condName == 'PE':
			df_cond1[df_cond.iat[k,0]] = bool(0 < df_cond.iat[k,colIndex]< 9)
		if condName == 'PB':
			df_cond1[df_cond.iat[k,0]] = bool(0 < df_cond.iat[k,colIndex]< 6)
		if condName == 'ROCE':
			df_cond1[df_cond.iat[k,0]] = bool(df_cond.iat[k,colIndex] > 15)
		if condName == 'DY':
			df_cond1[df_cond.iat[k,0]] = bool(df_cond.iat[k,colIndex] > 0.5)
		if condName == 'CR':
			df_cond1[df_cond.iat[k,0]] = bool(df_cond.iat[k,colIndex] > 2)
		if condName == 'DE':
			df_cond1[df_cond.iat[k,0]] = bool(df_cond.iat[k,colIndex] < 1)
		if condName == 'ROE':
			df_cond1[df_cond.iat[k,0]] = bool(df_cond.iat[k,colIndex] > 15)
	#print(df_cond1)
	cwd = os.path.abspath(os.getcwd())
	cwd = cwd.replace('\\', '/')
	cwd = cwd.replace('scripts', 'output')
	nameOfJson = cwd + '/' + condName + '.json'
	with open(nameOfJson, 'w') as fp:
		json.dump(df_cond1, fp)
	for k, v in df_cond1.items():
		if df_cond1[k] == True:
			listColCOND1.append(k)
	for k, v in df_cond1.items():
		if df_cond1[k] == False:
			listColCOND0.append(k)
	return df_cond1, listColCOND1, listColCOND0
#-----------------------------------------------------------------------------------------------------------------------------
def colorCode(listTrueVal, listFalseVal, targetCol, filename, paramName):
	'''
		Input:
			Condition True Values, Condition False Values, Condition Column Index, Filename, condition name
		Context:
			Color codes the excel 
	'''
	wb = load_workbook(filename) #Load the workbook
	ws = wb['Screener'] #Load the worksheet
	indexT = []
	#ws['B'] will return all cells on the B column until the last one (similar to max_row but it's only for the B column)
	for cell in ws['A']:
		if(cell.value is not None): #We need to check that the cell is not empty.
			if (cell.row == 1):
				continue
			elif (cell.value in listTrueVal): #Check if the value of the cell contains the text 'Table'
				ws.cell(row=cell.row, column=targetCol).fill = PatternFill(fgColor='008000', fill_type = 'solid')
			else:
				ws.cell(row=cell.row, column=targetCol).fill = PatternFill(fgColor='FF0000', fill_type = 'solid')
	wb.save(filename)
	print("Color coding for {} is done...".format(paramName))
#-----------------------------------------------------------------------------------------------------------------------------	
def getScreenerFile():
	'''
		Output:
			screener raw data dataframe
	'''
	cwd = os.path.abspath(os.getcwd())
	cwd = cwd.replace('\\', '/')
	cwd = cwd.replace('scripts', 'input')
	xlsxPath = cwd + '/*.csv'
	all_files = glob.glob(xlsxPath) #give path to your desired file path
	try:
		latest_csv = max(all_files, key=os.path.getctime)
		df = pd.read_csv(latest_csv)
	except ValueError:
		print("ValueError: " + f)
	print("File --> {} i.e.".format(latest_csv) + " is processed...")
	return df
#-----------------------------------------------------------------------------------------------------------------------------	
def priorityClassification(param0, param1, param2, param3, param4, param5, param6, param7):	
	cwd = os.path.abspath(os.getcwd())
	cwd = cwd.replace('\\', '/')
	cwd = cwd.replace('scripts', 'output')
	filename0 = cwd + '/dd.json'	
	dd = defaultdict(list)
	for d in (param0, param1, param2, param3, param4, param5, param6, param7): # you can list as many input dicts as you want here
		for key, value in d.items():
			dd[key].append(value)
	#print(dd)
	getDD = dict(dd)
	with open(filename0, 'w') as fp:
		json.dump(getDD, fp)
	dd1 = {}
	for k, v in getDD.items():
		numOfParams = sum(getDD[k])
		if numOfParams >= 8:
			dd1[k] = 'P0'
		elif 5 <= numOfParams <= 7:
			dd1[k] = 'P1'
		elif 3 <= numOfParams <= 4:
			dd1[k] = 'P2'
		else:
			dd1[k] = 'P3'
	#print(dd1)
	df_dd1 = pd.DataFrame(list(dd1.items()),columns = ['Name','Priority'])
	filename1 = cwd + '/dd1.json'
	#print(df_dd1)
	with open(filename1, 'w') as fp:
		json.dump(dd1, fp)
	return df_dd1
#-----------------------------------------------------------------------------------------------------------------------------	
start = time.time()
print("**************************************************** Hey You! Screener Script Started! ****************************************************")
screener_data = getScreenerFile()
screener_data.rename(columns = {'Market Capitalization':'MCAP', 'Price to Earning':'PE', 'Dividend yield':'DY', 'Price to book value':'PB', 'Current ratio':'CR', 'Debt to equity':'DE', 'Return on capital employed':'ROCE', 'Return on equity':'ROE' }, inplace = True)
#print(screener_data)
param0, listMC1, listMC0 = getCondScreener(screener_data, 'MCAP', screener_data.columns.get_loc("MCAP"))
param1, listPE1, listPE0 = getCondScreener(screener_data, 'PE', screener_data.columns.get_loc("PE"))
param2, listDY1, listDY0 = getCondScreener(screener_data, 'DY', screener_data.columns.get_loc("DY"))
param3, listPB1, listPB0 = getCondScreener(screener_data, 'PB', screener_data.columns.get_loc("PB"))
param4, listCR1, listCR0 = getCondScreener(screener_data, 'CR', screener_data.columns.get_loc("CR"))
param5, listDE1, listDE0 = getCondScreener(screener_data, 'DE', screener_data.columns.get_loc("DE"))
param6, listROCE1, listROCE0 = getCondScreener(screener_data, 'ROCE', screener_data.columns.get_loc("ROCE"))
param7, listROE1, listROE0 = getCondScreener(screener_data, 'ROE', screener_data.columns.get_loc("ROE"))
priority_data = priorityClassification(param0, param1, param2, param3, param4, param5, param6, param7)
priority_all_data = pd.merge(screener_data, priority_data, on = 'Name', how = 'left')
priority_all_data.rename(columns = {'Promoter holding':'Promoter Holding', 'MCAP':'Market Capitalization', 'PE':'Price To Earning', 'DY':'Dividend Yield', 'PB':'Price To Book Value', 'CR':'Current Ratio', 'DE':'Debt To Equity', 'ROCE':'Return On Capital Employed', 'ROE':'Return On Equity', 'Cash end of last year':'Cash At End Of Last Year', 'High price all time':'High Price All Time' }, inplace = True)
#print(list(priority_all_data.columns))
priority_all_data = priority_all_data.sort_values(by=['Priority'])
timestr = time.strftime("%Y%m%d_%H%M%S")
cwd = os.path.abspath(os.getcwd())
cwd = cwd.replace('\\', '/')
cwd = cwd.replace('scripts', 'output')
filename = cwd + '/screener_' + timestr + '.xlsx'
writer2 = pd.ExcelWriter(filename, engine='xlsxwriter')
priority_all_data.to_excel(writer2, sheet_name='Screener', index = False)
writer2.save()
print("Color decoded screener is saved...")
#print(priority_all_data.columns.get_loc("Market Capitalization"))
colorCode(listMC1, listMC0, priority_all_data.columns.get_loc("Market Capitalization") + 1, filename, 'Market Capitalization')
colorCode(listPE1, listPE0, priority_all_data.columns.get_loc("Price To Earning") + 1, filename, 'Price To Earning')
colorCode(listDY1, listDY0, priority_all_data.columns.get_loc("Dividend Yield") + 1, filename, 'Dividend Yield')
colorCode(listPB1, listPB0, priority_all_data.columns.get_loc("Price To Book Value") + 1, filename, 'Price To Book Value')
colorCode(listCR1, listCR0, priority_all_data.columns.get_loc("Current Ratio") + 1, filename, 'Current Ratio')
colorCode(listROCE1, listROCE0, priority_all_data.columns.get_loc("Return On Capital Employed") + 1, filename, 'Return On Capital Employed')
colorCode(listDE1, listDE0, priority_all_data.columns.get_loc("Debt To Equity") + 1, filename, 'Debt To Equity')
colorCode(listROE1, listROE0, priority_all_data.columns.get_loc("Return On Equity") + 1, filename, 'Return On Equity')
print("Color coding of screener is done...")
end = time.time()
hours, rem = divmod(end-start, 3600)
minutes, seconds = divmod(rem, 60)
print("Execution Time --> {:0>2}:{:0>2}:{:05.2f}".format(int(hours),int(minutes),seconds))
print("*********************************************************** See You Next Time!! ***********************************************************")
#-----------------------------------------------------------------------------------------------------------------------------